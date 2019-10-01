import os
import psycopg2
import xlrd
import xlwt
import openpyxl


class conprop:

    def __init__(self, host, port, db_name, user_name, password):

        self.host = host
        self.port = port
        self.db_name = db_name
        self.user_name = user_name
        self.password = password

    def constr(self):

        connect_str = "dbname='%s' user='%s' host='%s' port='%s' password='%s'" % (
            self.db_name, self.user_name, self.host, self.port, self.password)
        return connect_str

    def connect_db(self):

        try:
            connect_str = self.constr()
            conn = psycopg2.connect(connect_str)
            conn.autocommit = True
            self.cursor = conn.cursor()
            self.massage = 'Подключение установлено'
            self.load_table_settings()
            self.tb_name = str(self.tables[0])
            self.tb_name = self.tb_name.replace((','), (''))
            self.tb_name = self.tb_name.replace(('('), (''))
            self.tb_name = self.tb_name.replace((')'), (''))
            self.tb_name = self.tb_name.replace(("'"), (''))
            self.loaded_table_name(self.tb_name)

        except AttributeError as e:
            print("04")
            print(e)
        except Exception as e:
            self.massage = str(e)

    def load_table_settings(self):

        try:
            select_tables_name = "SELECT table_name FROM information_schema.tables WHERE table_schema='public'"
            self.cursor.execute(select_tables_name)
            self.tables = self.cursor.fetchall()
        except AttributeError as e:
            print("02")
            print(e)
            self.massage = 'Не было произведено подключение к БД'

    def loaded_table_name(self, tb_name):

        try:
            select_column_name = "SELECT * FROM %s" % (tb_name)
            self.tick_num = 0
            self.cursor.execute(select_column_name)
            column = [desc[0] for desc in self.cursor.description]
            self.column_str = ""
            self.column_param = ""
            for i in range(len(column)):
                self.column_str += "%s," % column[i]
            self.column_str = self.column_str.rstrip(',')
            self.massage = 'Данные о таблице %s успешно загруженны' % (tb_name)
            for i in range(0, len(column)):
                a = i+1
                self.column_param += "$%s," % a
            self.column_param = self.column_param.rstrip(',')
            self.tb_name = tb_name
            print("column param:  ", self.column_param)
        except AttributeError as e:
            print(e)
            print("03")
            self.massage = 'Не было произведено подключение к БД'
        except Exception as e:
            print("002")
            print(e)

    def load_data_to_server(self, values, numbers):

        self.tick_num += 1
        self.values_str = ""
        print("column param1:  ", self.column_param)
        self.prepare_string = "PREPARE trans%s AS INSERT INTO %s (%s) VALUES (%s);" % (
            self.tick_num, self.tb_name, self.column_str, self.column_param)
        print(self.prepare_string)
        self.cursor.execute(self.prepare_string)
        for i in range(len(values)):
            for j in numbers:
                self.values_str += "%s," % values[i][j-1]
            self.values_str = self.values_str.rstrip(',')
            print(self.values_str)
            execute_str = "execute trans%s (%s)" % (
                self.tick_num, self.values_str)
            self.cursor.execute(execute_str)
            self.values_str = ""
        print("trans num= ",self.tick_num)
        self.massage = 'Данные успешно загруженны на сервер'

class excel_edit:
    def __init__(self, file_name):

        self.file_name = file_name
        self.oef_pressed = False

    def read_excel(self):

        try:
            file_name_f, self.file_extention = os.path.splitext(self.file_name)
            if self.file_extention == '.xls':
                print("xls")
                self.excext = 0
                self.read_xls()
                self.massage = 'Успешно загружен файл Excel 97-2003'
                self.oef_pressed = True
            elif self.file_extention == '.xlsx':
                self.excext = 1
                print("xlsx")
                self.read_xlsx()
                self.massage = 'Успешно загружен файл Excel'
                self.oef_pressed = True
            print("ex= ", self.excext)

        except Exception as e:

            self.massage = str(e)
            self.oef_pressed = False

    def read_xlsx(self):

        self.wb = openpyxl.load_workbook(filename=self.file_name)
        print(self.wb.sheetnames)
        sheet = self.wb.active
        self.sheet_names = self.wb.sheetnames
        sheet = self.wb.active

    def read_xls(self):

        print("hz")
        self.rb = xlrd.open_workbook(self.file_name)
        self.sheet_names = self.rb.sheet_names()
        sheet = self.rb.sheet_by_index(0)
        self.vals = [sheet.row_values(rownum) for rownum in range(sheet.nrows)]

    def read_xls_sheet(self, sheet_name):

        self.sheet_name = sheet_name
        sheet = self.rb.sheet_by_name(sheet_name)
        self.vals = [sheet.row_values(rownum) for rownum in range(sheet.nrows)]

    def read_xlsx_sheet(self, sheet_name):

        self.sheet_name = sheet_name
        sheet = self.wb.active
        self.vals = sheet.rows
        print(self.vals)
        print(sheet)
        print(self.wb.sheetnames)

    def read_sheet(self, sheet_name):

        if self.excext == 0:
            print("xls 1")
            self.read_xls_sheet(sheet_name)
        else:
            print("xlsx 1")
            self.read_xlsx_sheet(sheet_name)

    def delete_spaces(self):

        num = 0
        lst = [0]*len(self.vals)
        for i in range(0, len(self.vals)):
            for j in range(0, len(self.vals[i])):
                if self.vals[i][j] == '':
                    lst[i] += 1
        for i in range(len(lst)):
            if lst[i] == 0:
                num += 1
        self.vals_1 = [0]*num
        for i in list(range(num)):
            self.vals_1[i] = [0]*len(self.vals[0])
        p = 0
        for i in range(0, len(self.vals)):
            if lst[i] == 0:
                for j in range(0, len(self.vals[i])):
                    self.vals_1[i-p][j] = self.vals[i][j]
            else:
                p += 1
        self.massage = 'Удалены строки с пробелами'

    def diapason(self, from_d, to):

        numbers = []
        if from_d == to:
            numbers.append(int(from_d))
        elif from_d < to:
            for i in range(from_d, to+1):
                numbers.append(i)
        elif from_d > to:
            for i in range(from_d, to-1, -1):
                numbers.append(i)
        self.numbers = numbers

    def pere(self, string):

        try:
            string = string.replace('.', ' ')
            string = string.replace(',', ' ')
            string = string.replace('/', ' ')
            string = string.replace('-', ' ')
            numbers = [int(n) for n in string.split()]
            self.numbers = numbers
        except Exception as e:
            print(e)

    def formating(self, numbers):

        self.value_str = ""
        for i in range(len(numbers)):
            self.value_str_custom += "$%s," % numbers[i]-1
        self.value_str_custom = self.value_str_custom.rstrip(',')

    def save_excel_xls(self, file_name_s):

        wb = xlwt.Workbook()
        ws = wb.add_sheet(self.sheet_name)
        print(self.vals_1)
        for i in range(len(self.vals_1)):
            for j in range(len(self.vals_1[0])):
                ws.write(i, j, self.vals_1[i][j])
        wb.save(file_name_s)

    def save_excel_xlsx(self, file_name_s):

        print('00000')

    def save_excel(self, file_name_s):

        self.file_name_s = file_name_s
        file_name_f, file_extention = os.path.splitext(file_name_s)
        print(self.file_extention)
        if file_extention == '.xls':
            self.save_excel_xls(file_name_s)
            self.massage = 'Успешно сохранен файл Excel 97-2003'
        elif file_extention == '.xlsx':
            self.save_excel_xlsx(file_name_s)
            self.massage = 'Успешно сохранен файл Excel'
        else:
            if self.file_extention == '.xls':
                self.save_excel_xls(file_name_s+'.xls')
                self.massage = 'Успешно сохранен файл Excel 97-2003'
            else:
                self.save_excel_xlsx(file_name_s+'.xlsx')
                self.massage = 'Успешно сохранен файл Excel'
            self.massage = 'Расширение не указано, взято от изначального файла'
